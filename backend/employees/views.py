from rest_framework import viewsets, filters
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from .models import Employee
from .serializers import EmployeeSerializer, EmployeeListSerializer

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class EmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.all().order_by('-created_at')
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['department', 'status', 'employee_type']
    search_fields = ['code', 'full_name', 'email', 'phone']

    def get_serializer_class(self):
        if self.action == 'list':
            return EmployeeListSerializer
        return EmployeeSerializer
    
    def perform_create(self, serializer):
        last = Employee.objects.order_by('-id').first()
        next_id = (last.id + 1) if last else 1
        code = f"NV{next_id:04d}"
        serializer.save(code=code)

    @action(detail=True, methods=['patch'])
    def deactivate(self, request, pk=None):
        employee = self.get_object()
        employee.status = 'inactive'
        employee.save()
        return Response({'message': f'{employee.full_name} đã ngưng hoạt động'})
    
    # Export Excel

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách nhân viên"

        # Header
        headers = ['Mã NV', 'Họ tên', 'Ngày sinh', 'Giới tính', 'CCCD',
                'Email', 'SĐT', 'Phòng ban', 'Chức vụ', 'Loại', 'Trạng thái', 'Ngày vào làm']
        
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        employees = self.get_queryset()
        for row, emp in enumerate(employees, 2):
            ws.cell(row=row, column=1, value=emp.code)
            ws.cell(row=row, column=2, value=emp.full_name)
            ws.cell(row=row, column=3, value=str(emp.date_of_birth))
            ws.cell(row=row, column=4, value='Nam' if emp.gender == 'M' else 'Nữ')
            ws.cell(row=row, column=5, value=emp.id_card)
            ws.cell(row=row, column=6, value=emp.email)
            ws.cell(row=row, column=7, value=emp.phone)
            ws.cell(row=row, column=8, value=emp.department.name if emp.department else '')
            ws.cell(row=row, column=9, value=emp.position.name if emp.position else '')
            ws.cell(row=row, column=10, value='Thử việc' if emp.employee_type == 'probation' else 'Chính thức')
            ws.cell(row=row, column=11, value='Đang làm' if emp.status == 'active' else 'Đã nghỉ')
            ws.cell(row=row, column=12, value=str(emp.start_date))

        # Auto width
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="nhan_vien.xlsx"'
        wb.save(response)
        return response